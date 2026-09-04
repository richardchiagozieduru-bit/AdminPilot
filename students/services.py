"""
Student services: admission numbering, single-student creation with its first
enrollment, class changes, the bulk-import parse/validate/commit pipeline, and
the plain-language activity timeline formatter.

Three rules from docs/02_Database.md shape everything here:

  * Admission numbers come from the per-institution counter ([Code]-[Year]-[Seq]),
    reserved with a row lock inside the same transaction as the student insert —
    never COUNT(*)/MAX(), both of which race and reuse numbers after a delete.
  * StudentEnrollment is append-only. Creating a student writes the first
    enrollment row; a class change writes another. Nothing updates an existing
    enrollment.
  * Nothing reaches the `students` table during a bulk import until commit — the
    upload only fills the BulkImportBatch/BulkImportRow staging tables.

The fee-assignment hook (`_assign_active_fee_structure`) imports billing inside
the function on purpose. docs/01_Architecture.md ADR-003 orders the apps
core <- academic <- students <- billing, and billing already does
`from students.models import Student` at module load. A module-level
`from billing.models import ...` here would close that into an import cycle, so
the import is deferred to call time, after every app is loaded — the same
house pattern core/forms.py uses for its receipt lookup. In Phase 4 no fee
structure has been created yet, so the hook is a no-op returning None; Phase 5
is where it starts finding a structure to assign.
"""

import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from django.db import transaction
from django.utils import timezone

from academic.models import Class, ClassStatus
from core.models import AuditLog, Institution, InstitutionNumberSequence
from core.services import format_sequence, next_sequence_number, write_audit_log

from .models import (
    BulkImportBatch,
    BulkImportRow,
    BulkImportStatus,
    Gender,
    RowValidationStatus,
    Student,
    StudentEnrollment,
)


class BulkImportError(Exception):
    """A workbook that cannot be processed at all — e.g. not a readable .xlsx.

    Row-level problems (a bad tab name, a missing field) are recorded on the
    staged row and surfaced in the preview, never raised. This is only for the
    file being unusable as a whole, which the upload view turns into a form
    error rather than a 500.
    """


# --------------------------------------------------------------------------- #
# Import template layout
# --------------------------------------------------------------------------- #
IMPORT_COLUMNS = (
    ("first_name", "First Name", True),
    ("middle_name", "Middle Name", False),
    ("last_name", "Last Name", True),
    ("gender", "Gender", True),
    ("date_of_birth", "Date of Birth", True),
    ("father_name", "Father's Name", False),
    ("mother_name", "Mother's Name", False),
    ("guardian_name", "Guardian Name", True),
    ("guardian_phone", "Guardian Phone", True),
    ("guardian_email", "Guardian Email", False),
    ("address", "Address", False),
)

IMPORT_HEADERS = [header for _, header, _ in IMPORT_COLUMNS]

_FIELD_LIMITS = {
    "first_name": 100,
    "middle_name": 100,
    "last_name": 100,
    "father_name": 100,
    "mother_name": 100,
    "guardian_name": 100,
    "guardian_phone": 20,
    "guardian_email": 254,
}

_ILLEGAL_SHEET_CHARS = "[]:*?/\\"


# --------------------------------------------------------------------------- #
# Admission numbering
# --------------------------------------------------------------------------- #
def _institution_year(institution):
    now = timezone.now().astimezone(ZoneInfo(institution.timezone))
    return now.year


def generate_admission_number(institution):
    year = _institution_year(institution)
    number = next_sequence_number(
        institution.pk, InstitutionNumberSequence.Kind.ADMISSION, year
    )
    return format_sequence(institution.code, year, number)


# --------------------------------------------------------------------------- #
# Fee-assignment hook
# --------------------------------------------------------------------------- #
def _assign_active_fee_structure(student, *, klass, session, term):
    from billing.models import FeeStructure, StudentFeeAssignment

    structure = (
        FeeStructure.unscoped.filter(
            institution_id=student.institution_id,
            klass=klass,
            session=session,
            term=term,
            is_active=True,
        )
        .order_by("-id")
        .first()
    )
    if structure is None:
        return None

    existing = StudentFeeAssignment.unscoped.filter(
        institution_id=student.institution_id,
        student=student,
        fee_structure=structure,
    ).first()
    if existing is not None:
        return existing

    from billing.services import _populate_student_fee_items

    assignment = StudentFeeAssignment.unscoped.create(
        institution_id=student.institution_id,
        student=student,
        fee_structure=structure,
        amount_due=structure.total_amount,
    )
    _populate_student_fee_items(assignment)
    return assignment



# --------------------------------------------------------------------------- #
# Single-student creation and class change
# --------------------------------------------------------------------------- #
@transaction.atomic
def create_student(
    *, institution, student, klass, session, term, actor=None, ip_address=None
):
    student.institution_id = institution.pk
    student.admission_number = generate_admission_number(institution)
    student.save()

    StudentEnrollment.unscoped.create(
        institution_id=institution.pk,
        student=student,
        klass=klass,
        session=session,
        term=term,
        enrolled_by=actor if getattr(actor, "pk", None) else None,
    )

    _assign_active_fee_structure(student, klass=klass, session=session, term=term)

    write_audit_log(
        institution_id=institution.pk,
        actor=actor,
        action="student.created",
        summary=f"Added student {student.full_name} ({student.admission_number})",
        target_type="Student",
        target_id=student.pk,
        detail={"admission_number": student.admission_number, "class": klass.name},
        ip_address=ip_address,
    )
    return student


@transaction.atomic
def change_student_class(
    *, student, klass, session, term, actor=None, ip_address=None
):
    enrollment = StudentEnrollment.unscoped.create(
        institution_id=student.institution_id,
        student=student,
        klass=klass,
        session=session,
        term=term,
        enrolled_by=actor if getattr(actor, "pk", None) else None,
    )

    _assign_active_fee_structure(student, klass=klass, session=session, term=term)

    write_audit_log(
        institution_id=student.institution_id,
        actor=actor,
        action="student.class_changed",
        summary=f"Moved {student.full_name} to {klass.name}",
        target_type="Student",
        target_id=student.pk,
        detail={"class": klass.name, "session": session.name, "term": term.name},
        ip_address=ip_address,
    )
    return enrollment


@transaction.atomic
def delete_student(
    *,
    student,
    keep_financial_records=False,
    actor,
    ip_address=None,
):
    """Delete a student.

    If keep_financial_records=False (clean plate):
      Purges all associated payment allocations, receipts, payments, credit transactions,
      fee assignments, student enrollments, and the student profile.
    If keep_financial_records=True:
      Soft-archives the student to inactive status so financial ledgers remain intact.
    Audit-logged.
    """
    student_pk = student.pk
    student_name = student.full_name
    admission_number = student.admission_number
    institution_id = student.institution_id

    if keep_financial_records:
        student.status = "inactive"
        student.save(update_fields=["status", "updated_at"])
        write_audit_log(
            institution_id=institution_id,
            actor=actor,
            action="student.archived",
            summary=f"Archived {student_name} ({admission_number}) to preserve financial records",
            target_type="Student",
            target_id=student_pk,
            ip_address=ip_address,
        )
        return {"action": "archived", "student_name": student_name}

    from billing.models import (
        PaymentItemAllocation,
        Payment,
        Receipt,
        CreditTransaction,
        StudentFeeAssignment,
    )

    # 1. Cleanly delete payment item allocations, receipts, payments for this student
    for assignment in student.fee_assignments.all():
        for payment in assignment.payments.all():
            PaymentItemAllocation.unscoped.filter(payment=payment).delete()
            CreditTransaction.unscoped.filter(source_payment=payment).delete()
            if hasattr(payment, "receipt") and payment.receipt:
                payment.receipt.delete()
            payment.delete()
        CreditTransaction.unscoped.filter(applied_to_assignment=assignment).delete()
        assignment.delete()

    # 2. Cleanly delete student enrollments
    student.enrollments.all().delete()

    # 3. Cleanly delete student profile
    student.delete()

    write_audit_log(
        institution_id=institution_id,
        actor=actor,
        action="student.deleted",
        summary=f"Deleted student {student_name} ({admission_number}) and purged all records",
        target_type="Student",
        target_id=student_pk,
        ip_address=ip_address,
    )
    return {"action": "deleted", "student_name": student_name}



# --------------------------------------------------------------------------- #
# Bulk import functions
# --------------------------------------------------------------------------- #
def _safe_sheet_title(name):
    for char in _ILLEGAL_SHEET_CHARS:
        name = name.replace(char, " ")
    return name.strip()[:31]


def _active_classes(institution_id):
    return list(
        Class.unscoped.filter(
            institution_id=institution_id, status=ClassStatus.ACTIVE
        ).order_by("order", "name")
    )


def build_import_template(institution):
    workbook = Workbook()
    workbook.remove(workbook.active)

    header_font = Font(bold=True)
    for klass in _active_classes(institution.pk):
        sheet = workbook.create_sheet(title=_safe_sheet_title(klass.name))
        sheet.append(IMPORT_HEADERS)
        for cell in sheet[1]:
            cell.font = header_font
        phone_col = IMPORT_HEADERS.index("Guardian Phone") + 1
        for row in range(1, 200):
            sheet.cell(row=row, column=phone_col).number_format = "@"
    return workbook


def _text(value, limit=None):
    if value is None:
        return ""
    text = str(value).strip()
    return text[:limit] if limit else text


def _normalize_gender(value):
    text = _text(value).casefold()
    if text in {"male", "m"}:
        return Gender.MALE
    if text in {"female", "f"}:
        return Gender.FEMALE
    return ""


def _coerce_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%m/%d/%Y"):
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _row_is_blank(values):
    return not any(_text(value) for value in values)


def _existing_student_keys(institution_id):
    return {
        (first.casefold(), last.casefold(), dob)
        for first, last, dob in Student.unscoped.filter(
            institution_id=institution_id
        ).values_list("first_name", "last_name", "date_of_birth")
    }


def _build_row(institution_id, batch, tab, row_number, values):
    fields = [field for field, _, _ in IMPORT_COLUMNS]
    data = dict(zip(fields, values))
    return BulkImportRow(
        institution_id=institution_id,
        batch=batch,
        sheet_name=tab,
        row_number=row_number,
        first_name=_text(data.get("first_name"), _FIELD_LIMITS["first_name"]),
        middle_name=_text(data.get("middle_name"), _FIELD_LIMITS["middle_name"]),
        last_name=_text(data.get("last_name"), _FIELD_LIMITS["last_name"]),
        gender=_normalize_gender(data.get("gender")),
        date_of_birth=_coerce_date(data.get("date_of_birth")),
        father_name=_text(data.get("father_name"), _FIELD_LIMITS["father_name"]),
        mother_name=_text(data.get("mother_name"), _FIELD_LIMITS["mother_name"]),
        guardian_name=_text(data.get("guardian_name"), _FIELD_LIMITS["guardian_name"]),
        guardian_phone=_text(
            data.get("guardian_phone"), _FIELD_LIMITS["guardian_phone"]
        ),
        guardian_email=_text(
            data.get("guardian_email"), _FIELD_LIMITS["guardian_email"]
        ),
        address=_text(data.get("address")),
    )


def _row_error(row, seen, existing):
    missing = [
        label
        for field, label, required in IMPORT_COLUMNS
        if required and field != "gender" and not getattr(row, field)
    ]
    if missing:
        return "Missing " + ", ".join(missing) + "."

    if row.gender not in Gender.values:
        return "Gender must be “Male” or “Female”."

    key = (row.first_name.casefold(), row.last_name.casefold(), row.date_of_birth)
    if key in seen:
        return "Another row in this file has the same name and date of birth."
    if key in existing:
        return "A student with this name and date of birth already exists."

    seen.add(key)
    return ""


def _validate_row(row, *, klass, seen, existing):
    if klass is None:
        error = (
            f"Worksheet “{row.sheet_name}” does not match an active class."
        )
    else:
        error = _row_error(row, seen, existing)

    if error:
        row.validation_status = RowValidationStatus.ERROR
        row.error_reason = error
    else:
        row.validation_status = RowValidationStatus.VALID
        row.error_reason = ""


@transaction.atomic
def stage_upload(
    *, institution, uploaded_file, session, term, uploaded_by, ip_address=None
):
    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:
        raise BulkImportError(
            "That file could not be read as an Excel .xlsx workbook. Download the "
            "template, fill it in, and upload that file."
        ) from exc

    class_by_title = {
        _safe_sheet_title(klass.name).casefold(): klass
        for klass in _active_classes(institution.pk)
    }

    batch = BulkImportBatch.unscoped.create(
        institution_id=institution.pk,
        uploaded_by=uploaded_by if getattr(uploaded_by, "pk", None) else None,
        session=session,
        term=term,
        status=BulkImportStatus.PARSING,
    )

    seen = set()
    existing = _existing_student_keys(institution.pk)

    for sheet in workbook.worksheets:
        klass = class_by_title.get(_safe_sheet_title(sheet.title).casefold())
        rows = sheet.iter_rows(min_row=2, values_only=True)
        for offset, values in enumerate(rows, start=2):
            if _row_is_blank(values):
                continue
            row = _build_row(institution.pk, batch, sheet.title, offset, values)
            _validate_row(row, klass=klass, seen=seen, existing=existing)
            row.save()

    workbook.close()

    batch.status = BulkImportStatus.READY
    batch.save(update_fields=["status"])
    return batch


@transaction.atomic
def commit_import(*, batch, actor=None, ip_address=None):
    if batch.status == BulkImportStatus.COMMITTED:
        return 0

    institution = Institution.objects.get(pk=batch.institution_id)
    class_by_title = {
        _safe_sheet_title(klass.name).casefold(): klass
        for klass in _active_classes(institution.pk)
    }

    rows = list(
        BulkImportRow.unscoped.filter(
            batch=batch,
            validation_status=RowValidationStatus.VALID,
            selected_for_commit=True,
        )
    )

    created = 0
    for row in rows:
        klass = class_by_title.get(_safe_sheet_title(row.sheet_name).casefold())
        if klass is None:
            continue

        student = Student.unscoped.create(
            institution_id=institution.pk,
            first_name=row.first_name,
            middle_name=row.middle_name,
            last_name=row.last_name,
            gender=row.gender,
            date_of_birth=row.date_of_birth,
            admission_number=generate_admission_number(institution),
            father_name=row.father_name,
            mother_name=row.mother_name,
            guardian_name=row.guardian_name,
            guardian_phone=row.guardian_phone,
            guardian_email=row.guardian_email,
            address=row.address,
        )
        StudentEnrollment.unscoped.create(
            institution_id=institution.pk,
            student=student,
            klass=klass,
            session=batch.session,
            term=batch.term,
            enrolled_by=actor if getattr(actor, "pk", None) else None,
        )
        _assign_active_fee_structure(
            student, klass=klass, session=batch.session, term=batch.term
        )
        created += 1

    write_audit_log(
        institution_id=institution.pk,
        actor=actor,
        action="student.bulk_imported",
        summary=(
            f"Bulk imported {created} student{'' if created == 1 else 's'} "
            f"for {batch.session.name} — {batch.term.name}"
        ),
        target_type="BulkImportBatch",
        target_id=batch.pk,
        detail={
            "created": created,
            "session": batch.session.name,
            "term": batch.term.name,
        },
        ip_address=ip_address,
    )

    batch.status = BulkImportStatus.COMMITTED
    batch.save(update_fields=["status"])
    BulkImportRow.unscoped.filter(batch=batch).delete()
    return created


# --------------------------------------------------------------------------- #
# Activity Timeline
# --------------------------------------------------------------------------- #
def get_student_timeline(student):
    """Compile audit logs, enrollments, and payments for `student` into plain-language events."""
    events = []

    # 1. Audit logs
    audit_logs = AuditLog.unscoped.filter(
        institution_id=student.institution_id,
        target_type="Student",
        target_id=str(student.pk),
    ).select_related("actor")

    for log in audit_logs:
        actor_name = log.actor.full_name if log.actor else "System"
        events.append({
            "timestamp": log.created_at,
            "title": log.action.replace(".", " ").title(),
            "summary": f"{log.summary} by {actor_name}",
            "badge_class": "badge--info",
        })

    # 2. Enrollments
    enrollments = StudentEnrollment.unscoped.filter(
        student=student
    ).select_related("klass", "session", "term", "enrolled_by")

    for e in enrollments:
        actor_name = e.enrolled_by.full_name if e.enrolled_by else "System"
        events.append({
            "timestamp": e.enrolled_at,
            "title": "Enrolled into class",
            "summary": f"Enrolled into {e.klass.name} ({e.session.name} — {e.term.name}) by {actor_name}",
            "badge_class": "badge--success",
        })

    # 3. Payments
    from billing.models import Payment
    payments = Payment.unscoped.filter(
        assignment__student=student
    ).select_related("assignment__fee_structure", "receipt", "recorded_by")

    for p in payments:
        receipt_num = p.receipt.receipt_number if hasattr(p, "receipt") and p.receipt else "N/A"
        actor_name = p.recorded_by.full_name if p.recorded_by else "System"
        status_str = f" ({p.get_status_display()})" if p.status != "active" else ""
        events.append({
            "timestamp": p.created_at,
            "title": f"Payment Recorded{status_str}",
            "summary": f"Payment of {p.amount} ({p.get_method_display()}) recorded for {p.assignment.fee_structure.name} — Receipt {receipt_num} by {actor_name}",
            "badge_class": "badge--danger" if p.status == "reversed" else "badge--success",
        })

    events.sort(key=lambda x: x["timestamp"], reverse=True)
    return events
