"""
Phase 4 — Student Management. The exit condition (docs/07_Implementation_Roadmap.md):
a student can be added one at a time and via the full template → upload → preview
→ commit bulk import, and found again through the list's filters. These tests
drive every one of those paths through a real request.

`self.in_school()` wraps every assertion that reads or writes a tenant-scoped
table. A test method is not a request, so TenantContextMiddleware has stamped no
SESSION_CONTEXT — an unwrapped `Student.objects.count()` reads through RLS as
zero and the assertion fails for the wrong reason. During `self.client.*` the
middleware handles the stamp; the wrapper is only for the test body around it.
See core/tests/school.py.

Three things carry their own weight beyond ordinary CRUD, and each has a test
that fails if it is softened:

  * Class lives on the enrollment, not the student — adding or moving a student
    writes an append-only StudentEnrollment row, never an UPDATE.
  * A Bursar reads students but cannot manage them, and on the profile sees a
    field-restricted subset (docs/04). The manage refusal happens before the
    view body (RoleRequiredMixin), so an unauthorised POST must leave no row.
  * Nothing reaches the `students` table during a bulk import until commit — the
    upload only fills the staging tables, and commit is the single writer.
"""

import datetime
import re
from io import BytesIO

from openpyxl import Workbook, load_workbook

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse

from academic.models import Class, ClassStatus, Session, Term
from accounts.models import User
from core.middleware import institution_db_context
from core.models import AuditLog, Institution
from core.tests.school import (
    SESSION_END,
    SESSION_START,
    TERM_END,
    TERM_START,
    ApprovedSchoolTestCase,
)

from students.imports import XLSX_CONTENT_TYPE
from students.models import (
    BulkImportBatch,
    BulkImportRow,
    BulkImportStatus,
    Gender,
    RowValidationStatus,
    Student,
    StudentEnrollment,
    StudentStatus,
)
from students.services import IMPORT_HEADERS, delete_student


def other_school():
    """A second tenant. Institution carries no institution_id, so it is not
    behind the RLS predicate and needs no context to create."""
    return Institution.objects.create(
        name="Beta School", code="BETA", status=Institution.Status.APPROVED
    )


def make_workbook(sheets):
    """An in-memory .xlsx mirroring the import template: a header row per sheet
    followed by the given rows.

    `sheets` maps a worksheet title (which doubles as the class/tab name) to a
    list of rows, each a list of values in IMPORT_HEADERS order. Returns the
    workbook bytes, ready to wrap in a SimpleUploadedFile.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)  # drop the default empty sheet
    for title, rows in sheets.items():
        sheet = workbook.create_sheet(title=title)
        sheet.append(list(IMPORT_HEADERS))
        for row_values in rows:
            sheet.append(list(row_values))
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def sheet_row(
    first,
    last,
    *,
    gender="Male",
    dob="2012-05-01",
    middle="",
    father="",
    mother="",
    guardian="Ada Guardian",
    phone="08030000000",
    email="",
    address="",
):
    """One template row in column order (IMPORT_HEADERS): the order the parser
    reads positionally, so the tuple here is what a filled-in template looks
    like on the wire."""
    return [
        first,
        middle,
        last,
        gender,
        dob,
        father,
        mother,
        guardian,
        phone,
        email,
        address,
    ]


class StudentsTestCase(ApprovedSchoolTestCase):
    """A configured school (session, term, JSS 1A/JSS 1B) plus the helpers the
    student suites share. Concrete cases call super().setUp() and sign in."""

    def setUp(self):
        self.session, self.term, self.classes = self.configure_school()
        self.jss1a, self.jss1b = self.classes

    # -- seeding ----------------------------------------------------------- #
    def make_student(
        self,
        first,
        last,
        klass,
        *,
        suffix,
        status=StudentStatus.ACTIVE,
        gender=Gender.MALE,
        dob=datetime.date(2012, 4, 2),
        guardian_name="Ngozi Okeke",
        guardian_phone="08010000000",
        **extra,
    ):
        """A saved, enrolled student created straight through `unscoped`.

        Bypasses the create service on purpose so list/profile/edit tests start
        from a known row without spending the admission counter — `suffix` fixes
        the admission number, so callers control uniqueness. Tests that mean to
        exercise the counter go through the add view instead.
        """
        with self.in_school():
            student = Student.unscoped.create(
                institution=self.institution,
                first_name=first,
                last_name=last,
                gender=gender,
                date_of_birth=dob,
                admission_number=f"{self.institution.code}-2026-{suffix}",
                guardian_name=guardian_name,
                guardian_phone=guardian_phone,
                status=status,
                **extra,
            )
            StudentEnrollment.unscoped.create(
                institution=self.institution,
                student=student,
                klass=klass,
                session=self.session,
                term=self.term,
            )
        return student

    # -- POST payloads ----------------------------------------------------- #
    def add_post(self, klass_pk, **overrides):
        """A complete, valid add-student POST. Overrides tweak one field."""
        data = {
            "first_name": "Amaka",
            "middle_name": "",
            "last_name": "Eze",
            "gender": Gender.FEMALE,
            "date_of_birth": "2013-03-04",
            "klass": klass_pk,
            "guardian_name": "Ada Eze",
            "guardian_phone": "08030000000",
            "guardian_email": "",
            "father_name": "",
            "mother_name": "",
            "address": "",
            "date_of_admission": "",
        }
        data.update(overrides)
        return data

    def edit_post(self, student, klass_pk, **overrides):
        """The edit form pre-filled from the row, so `changed_data` is accurate.

        Literals would differ from the stored student in fields nobody touched,
        and a "saved with no change" test would then be testing the opposite —
        the same reasoning the settings fixture uses in core/tests/school.py.
        """
        data = {
            "first_name": student.first_name,
            "middle_name": student.middle_name,
            "last_name": student.last_name,
            "gender": student.gender,
            "date_of_birth": student.date_of_birth.isoformat(),
            "klass": klass_pk,
            "guardian_name": student.guardian_name,
            "guardian_phone": student.guardian_phone,
            "guardian_email": student.guardian_email,
            "father_name": student.father_name,
            "mother_name": student.mother_name,
            "address": student.address,
            "date_of_admission": (
                student.date_of_admission.isoformat()
                if student.date_of_admission
                else ""
            ),
        }
        data.update(overrides)
        return data

    # -- bulk import ------------------------------------------------------- #
    def upload(self, sheets, term=None):
        """POST a generated workbook to the upload view. Caller signs in first."""
        upload = SimpleUploadedFile(
            "import.xlsx", make_workbook(sheets), content_type=XLSX_CONTENT_TYPE
        )
        return self.client.post(
            reverse("students:import"),
            {"term": (term or self.term).pk, "file": upload},
        )

    def latest_batch(self):
        with self.in_school():
            return BulkImportBatch.objects.latest("pk")

    def valid_row_pks(self, batch):
        with self.in_school():
            return list(
                BulkImportRow.objects.filter(
                    batch=batch, validation_status=RowValidationStatus.VALID
                )
                .order_by("pk")
                .values_list("pk", flat=True)
            )

    def all_row_pks(self, batch):
        with self.in_school():
            return list(
                BulkImportRow.objects.filter(batch=batch)
                .order_by("pk")
                .values_list("pk", flat=True)
            )

    def commit(self, batch, row_pks):
        """POST the preview, selecting `row_pks`. An empty list selects nobody."""
        return self.client.post(
            reverse("students:import_preview", args=[batch.pk]),
            {"rows": [str(pk) for pk in row_pks]},
        )


# --------------------------------------------------------------------------- #
# List + filters
# --------------------------------------------------------------------------- #
class StudentListTests(StudentsTestCase):
    def setUp(self):
        super().setUp()
        self.url = reverse("students:list")

    def test_the_list_shows_each_students_current_class(self):
        self.make_student("Amaka", "Eze", self.jss1a, suffix="000001")
        self.make_student("Bola", "Okoro", self.jss1b, suffix="000002")
        self.sign_in_owner()

        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        classes = {
            student.full_name: student.current_class_name
            for student in response.context["students"]
        }
        self.assertEqual(classes, {"Amaka Eze": "JSS 1A", "Bola Okoro": "JSS 1B"})

    def test_the_list_hides_archived_students_until_asked(self):
        """Default is active-only — an archived student is kept, not shown, and
        the status filter is what brings them back."""
        self.make_student("Active", "One", self.jss1a, suffix="000001")
        self.make_student(
            "Archived", "Two", self.jss1a, suffix="000002", status=StudentStatus.INACTIVE
        )
        self.sign_in_owner()

        default = self.client.get(self.url)
        self.assertEqual([s.full_name for s in default.context["students"]], ["Active One"])

        archived = self.client.get(self.url, {"status": "inactive"})
        self.assertEqual(
            [s.full_name for s in archived.context["students"]], ["Archived Two"]
        )

        every = self.client.get(self.url, {"status": "all"})
        self.assertEqual(len(every.context["students"]), 2)

    def test_filtering_by_class_uses_the_current_enrollment(self):
        self.make_student("Amaka", "Eze", self.jss1a, suffix="000001")
        self.make_student("Bola", "Okoro", self.jss1b, suffix="000002")
        self.sign_in_owner()

        response = self.client.get(self.url, {"class_id": self.jss1a.pk})
        self.assertEqual(
            [s.full_name for s in response.context["students"]], ["Amaka Eze"]
        )

    def test_search_matches_name_or_admission_number(self):
        self.make_student("Amaka", "Eze", self.jss1a, suffix="000001")
        bola = self.make_student("Bola", "Okoro", self.jss1b, suffix="000002")
        self.sign_in_owner()

        by_name = self.client.get(self.url, {"q": "okoro"})
        self.assertEqual([s.full_name for s in by_name.context["students"]], ["Bola Okoro"])

        by_admission = self.client.get(self.url, {"q": bola.admission_number})
        self.assertEqual(
            [s.full_name for s in by_admission.context["students"]], ["Bola Okoro"]
        )

    def test_another_schools_students_are_not_listed(self):
        """The tenant filter through a real request — the application layer
        agreeing with RLS, and the layer a mistake is likeliest to slip into."""
        self.make_student("Ours", "Here", self.jss1a, suffix="000001")
        other = other_school()
        with institution_db_context(other.pk):
            Student.unscoped.create(
                institution=other,
                first_name="Beta",
                last_name="Student",
                gender=Gender.MALE,
                date_of_birth=datetime.date(2012, 1, 1),
                admission_number="BETA-2026-000001",
                guardian_name="G",
                guardian_phone="0",
            )

        self.sign_in_owner()
        response = self.client.get(self.url)
        self.assertNotContains(response, "Beta Student")
        self.assertEqual(len(response.context["students"]), 1)

    def test_a_bursar_reads_the_list_without_manage_controls(self):
        """students is view=FULL, manage=Owner/Admin. The add/import controls are
        absent from a Bursar's markup, not disabled — no path they'd be refused on."""
        self.make_student("Amaka", "Eze", self.jss1a, suffix="000001")
        self.sign_in_as(User.Role.BURSAR)

        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["can_manage"])
        self.assertContains(response, "Amaka Eze")
        self.assertNotContains(response, reverse("students:add"))
        self.assertNotContains(response, reverse("students:import"))

    def test_a_staff_account_reaches_nothing(self):
        """Staff is inert in V1 — in no view set of the matrix."""
        self.sign_in_as(User.Role.STAFF)
        self.assertEqual(self.client.get(self.url).status_code, 403)


# --------------------------------------------------------------------------- #
# Create
# --------------------------------------------------------------------------- #
class StudentCreateTests(StudentsTestCase):
    def setUp(self):
        super().setUp()
        self.url = reverse("students:add")

    def test_adding_a_student_writes_row_enrollment_number_and_audit(self):
        self.sign_in_owner()
        response = self.client.post(self.url, self.add_post(self.jss1a.pk))

        with self.in_school():
            student = Student.objects.get(first_name="Amaka")
        self.assertRedirects(response, reverse("students:detail", args=[student.pk]))

        with self.in_school():
            self.assertEqual(student.institution_id, self.institution.pk)
            self.assertRegex(
                student.admission_number,
                rf"^{re.escape(self.institution.code)}-\d{{4}}-\d{{6}}$",
            )
            enrollment = StudentEnrollment.objects.get(student=student)
            self.assertEqual(enrollment.klass_id, self.jss1a.pk)
            self.assertEqual(enrollment.session_id, self.session.pk)
            self.assertEqual(enrollment.term_id, self.term.pk)
            self.assertEqual(enrollment.enrolled_by.email, self.OWNER_EMAIL)

            entry = AuditLog.objects.filter(action="student.created").get()
        self.assertEqual(entry.target_id, str(student.pk))
        self.assertEqual(entry.detail["class"], "JSS 1A")
        self.assertEqual(entry.detail["admission_number"], student.admission_number)
        self.assertIn(student.admission_number, entry.summary)

    def test_admission_numbers_increment_per_institution(self):
        """Sequential from the counter, not COUNT(*)/MAX — reserved with a row
        lock inside the insert transaction, so two adds never collide or reuse."""
        self.sign_in_owner()
        self.client.post(self.url, self.add_post(self.jss1a.pk, first_name="Amaka"))
        self.client.post(self.url, self.add_post(self.jss1a.pk, first_name="Bola"))

        with self.in_school():
            numbers = sorted(Student.objects.values_list("admission_number", flat=True))
        self.assertTrue(numbers[0].endswith("-000001"), numbers)
        self.assertTrue(numbers[1].endswith("-000002"), numbers)

    def test_adding_needs_a_current_term_to_enrol_into(self):
        """No current term means no term to enroll into. The view redirects to
        set one rather than failing at save — and creates nothing meanwhile."""
        with self.in_school():
            Session.unscoped.filter(institution_id=self.institution.pk).update(
                is_current=False
            )
            Term.unscoped.filter(institution_id=self.institution.pk).update(
                is_current=False
            )
        self.sign_in_owner()

        get_response = self.client.get(self.url)
        self.assertEqual(get_response.status_code, 302)
        self.assertIn(reverse("academic:structure"), get_response["Location"])

        post_response = self.client.post(self.url, self.add_post(self.jss1a.pk))
        self.assertEqual(post_response.status_code, 302)
        with self.in_school():
            self.assertEqual(Student.objects.count(), 0)

    def test_a_bursar_post_creates_nothing(self):
        """The ordering RoleRequiredMixin.check_access exists for: a manage check
        that ran after the body would insert the row and then return 403."""
        self.sign_in_as(User.Role.BURSAR)
        response = self.client.post(self.url, self.add_post(self.jss1a.pk))

        self.assertEqual(response.status_code, 403)
        with self.in_school():
            self.assertEqual(
                Student.objects.count(),
                0,
                "A Bursar's POST created a student before being refused.",
            )

    def test_another_schools_class_is_refused_as_a_field_error(self):
        """A hand-built POST naming a foreign class id has to come back as a
        field error, not a student filed under someone else's class."""
        other = other_school()
        with institution_db_context(other.pk):
            foreign = Class.unscoped.create(
                institution=other, name="Beta Only", order=10
            )
        self.sign_in_owner()

        response = self.client.post(self.url, self.add_post(foreign.pk))
        self.assertEqual(response.status_code, 200)
        self.assertIn("klass", response.context["form"].errors)
        with self.in_school():
            self.assertEqual(Student.objects.count(), 0)


# --------------------------------------------------------------------------- #
# Profile (tabs + Bursar field restriction)
# --------------------------------------------------------------------------- #
class StudentProfileTests(StudentsTestCase):
    def setUp(self):
        super().setUp()
        self.student = self.make_student(
            "Amaka",
            "Eze",
            self.jss1a,
            suffix="000001",
            guardian_email="secret@example.com",
            father_name="Papa Secret",
            mother_name="Mama Secret",
            address="12 Secret Road",
        )

    def test_owner_overview_shows_the_full_record(self):
        self.sign_in_owner()
        response = self.client.get(reverse("students:detail", args=[self.student.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["is_restricted"])
        labels = [label for label, _ in response.context["overview_rows"]]
        for expected in ("Date of birth", "Address", "Guardian email", "Father's name"):
            self.assertIn(expected, labels)
        self.assertContains(response, "12 Secret Road")

    def test_a_bursar_overview_hides_the_restricted_fields(self):
        """docs/04: a Bursar sees class, primary guardian contact and balance —
        not DOB, address, email or parent names. Enforced in the context the view
        builds, so the restricted fields never reach the template to be leaked."""
        self.sign_in_as(User.Role.BURSAR)
        response = self.client.get(reverse("students:detail", args=[self.student.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["is_restricted"])
        labels = [label for label, _ in response.context["overview_rows"]]
        self.assertEqual(
            labels, ["Class", "Guardian name", "Guardian phone", "Credit balance"]
        )
        # The restricted values are absent from the page, not merely unlabelled.
        self.assertNotContains(response, "12 Secret Road")
        self.assertNotContains(response, "Papa Secret")
        self.assertNotContains(response, "secret@example.com")
        self.assertNotContains(response, "Date of birth")
        # ...while the fields a Bursar is entitled to are present.
        self.assertContains(response, self.student.guardian_phone)

    def test_the_enrolment_tab_lists_history_newest_first(self):
        """A class change appends a row; the tab shows both, latest first, so the
        append-only history stays visible rather than being overwritten."""
        self.sign_in_owner()
        self.client.post(
            reverse("students:edit", args=[self.student.pk]),
            self.edit_post(self.student, self.jss1b.pk),
        )

        response = self.client.get(
            reverse("students:enrollments", args=[self.student.pk])
        )
        self.assertEqual(
            [e.klass.name for e in response.context["enrollments"]],
            ["JSS 1B", "JSS 1A"],
        )

    def test_payment_and_timeline_tabs_render(self):
        self.sign_in_owner()
        payments = self.client.get(reverse("students:payments", args=[self.student.pk]))
        self.assertEqual(payments.status_code, 200)
        self.assertContains(payments, "Fee Packages")
        timeline = self.client.get(reverse("students:timeline", args=[self.student.pk]))
        self.assertEqual(timeline.status_code, 200)

    def test_another_schools_student_profile_is_a_404(self):
        """Scoped fetch → the student does not exist for this tenant. A 404, not
        a 403: a 403 would confirm the id is real."""
        other = other_school()
        with institution_db_context(other.pk):
            foreign = Student.unscoped.create(
                institution=other,
                first_name="Beta",
                last_name="Student",
                gender=Gender.MALE,
                date_of_birth=datetime.date(2012, 1, 1),
                admission_number="BETA-2026-000001",
                guardian_name="G",
                guardian_phone="0",
            )
        self.sign_in_owner()
        response = self.client.get(reverse("students:detail", args=[foreign.pk]))
        self.assertEqual(response.status_code, 404)


# --------------------------------------------------------------------------- #
# Edit
# --------------------------------------------------------------------------- #
class StudentUpdateTests(StudentsTestCase):
    def setUp(self):
        super().setUp()
        self.student = self.make_student("Amaka", "Eze", self.jss1a, suffix="000001")

    def test_editing_biography_saves_and_logs_the_changed_fields(self):
        self.sign_in_owner()
        response = self.client.post(
            reverse("students:edit", args=[self.student.pk]),
            self.edit_post(self.student, self.jss1a.pk, guardian_phone="09099999999"),
        )
        self.assertRedirects(response, reverse("students:detail", args=[self.student.pk]))

        with self.in_school():
            self.student.refresh_from_db()
            self.assertEqual(self.student.guardian_phone, "09099999999")
            # No class change → still exactly one enrollment.
            self.assertEqual(
                StudentEnrollment.objects.filter(student=self.student).count(), 1
            )
            entry = AuditLog.objects.filter(action="student.updated").get()
            self.assertFalse(
                AuditLog.objects.filter(action="student.class_changed").exists()
            )
        self.assertEqual(entry.detail["fields"], ["guardian_phone"])

    def test_changing_class_appends_an_enrollment_and_logs_it(self):
        """The class move is a new StudentEnrollment, never an update — the first
        row survives, the new one becomes current, and the change is audited."""
        self.sign_in_owner()
        self.client.post(
            reverse("students:edit", args=[self.student.pk]),
            self.edit_post(self.student, self.jss1b.pk),
        )

        with self.in_school():
            enrollments = list(
                StudentEnrollment.objects.filter(student=self.student).order_by(
                    "enrolled_at"
                )
            )
            self.assertEqual([e.klass_id for e in enrollments], [self.jss1a.pk, self.jss1b.pk])
            self.assertTrue(
                AuditLog.objects.filter(action="student.class_changed").exists()
            )
            # Bio untouched, so no student.updated entry rode along.
            self.assertFalse(
                AuditLog.objects.filter(action="student.updated").exists()
            )

    def test_saving_with_no_changes_writes_no_audit_and_no_enrollment(self):
        self.sign_in_owner()
        self.client.post(
            reverse("students:edit", args=[self.student.pk]),
            self.edit_post(self.student, self.jss1a.pk),
        )
        with self.in_school():
            self.assertEqual(
                StudentEnrollment.objects.filter(student=self.student).count(), 1
            )
            self.assertFalse(AuditLog.objects.filter(action="student.updated").exists())
            self.assertFalse(
                AuditLog.objects.filter(action="student.class_changed").exists()
            )

    def test_a_bursar_cannot_edit(self):
        self.sign_in_as(User.Role.BURSAR)
        response = self.client.post(
            reverse("students:edit", args=[self.student.pk]),
            self.edit_post(self.student, self.jss1a.pk, guardian_phone="09099999999"),
        )
        self.assertEqual(response.status_code, 403)
        with self.in_school():
            self.student.refresh_from_db()
        self.assertEqual(self.student.guardian_phone, "08010000000")

    def test_editing_another_schools_student_is_a_404(self):
        other = other_school()
        with institution_db_context(other.pk):
            foreign = Student.unscoped.create(
                institution=other,
                first_name="Beta",
                last_name="Student",
                gender=Gender.MALE,
                date_of_birth=datetime.date(2012, 1, 1),
                admission_number="BETA-2026-000001",
                guardian_name="G",
                guardian_phone="0",
            )
        self.sign_in_owner()
        response = self.client.post(
            reverse("students:edit", args=[foreign.pk]),
            self.add_post(self.jss1a.pk),
        )
        self.assertEqual(response.status_code, 404)


# --------------------------------------------------------------------------- #
# Archive / reactivate
# --------------------------------------------------------------------------- #
class StudentArchiveReactivateTests(StudentsTestCase):
    def setUp(self):
        super().setUp()
        self.student = self.make_student("Amaka", "Eze", self.jss1a, suffix="000001")

    def test_archiving_is_a_soft_delete_that_keeps_the_history(self):
        self.sign_in_owner()
        response = self.client.post(reverse("students:archive", args=[self.student.pk]))
        self.assertRedirects(response, reverse("students:detail", args=[self.student.pk]))

        with self.in_school():
            self.student.refresh_from_db()
            self.assertEqual(self.student.status, StudentStatus.INACTIVE)
            # The enrollment row is untouched — a soft delete moves only status.
            self.assertEqual(
                StudentEnrollment.objects.filter(student=self.student).count(), 1
            )
            entry = AuditLog.objects.filter(action="student.archived").get()
        self.assertEqual(entry.target_id, str(self.student.pk))

    def test_archiving_is_post_only(self):
        """A state change reachable by GET can be fired by a link prefetch."""
        self.sign_in_owner()
        response = self.client.get(reverse("students:archive", args=[self.student.pk]))
        self.assertEqual(response.status_code, 405)
        with self.in_school():
            self.student.refresh_from_db()
        self.assertEqual(self.student.status, StudentStatus.ACTIVE)

    def test_reactivating_restores_an_archived_student(self):
        self.sign_in_owner()
        self.client.post(reverse("students:archive", args=[self.student.pk]))
        self.client.post(reverse("students:reactivate", args=[self.student.pk]))

        with self.in_school():
            self.student.refresh_from_db()
            self.assertEqual(self.student.status, StudentStatus.ACTIVE)
            self.assertTrue(
                AuditLog.objects.filter(action="student.reactivated").exists()
            )

    def test_archiving_twice_is_a_no_op_not_a_second_log_entry(self):
        self.sign_in_owner()
        url = reverse("students:archive", args=[self.student.pk])
        self.client.post(url)
        self.client.post(url)
        with self.in_school():
            self.assertEqual(
                AuditLog.objects.filter(action="student.archived").count(),
                1,
                "A repeated archive POST logged a second archival of an "
                "already-archived student.",
            )

    def test_a_bursar_cannot_archive(self):
        self.sign_in_as(User.Role.BURSAR)
        response = self.client.post(reverse("students:archive", args=[self.student.pk]))
        self.assertEqual(response.status_code, 403)
        with self.in_school():
            self.student.refresh_from_db()
        self.assertEqual(self.student.status, StudentStatus.ACTIVE)

    def test_archiving_another_schools_student_is_a_404(self):
        other = other_school()
        with institution_db_context(other.pk):
            foreign = Student.unscoped.create(
                institution=other,
                first_name="Beta",
                last_name="Student",
                gender=Gender.MALE,
                date_of_birth=datetime.date(2012, 1, 1),
                admission_number="BETA-2026-000001",
                guardian_name="G",
                guardian_phone="0",
            )
        self.sign_in_owner()
        response = self.client.post(reverse("students:archive", args=[foreign.pk]))
        self.assertEqual(response.status_code, 404)
        with institution_db_context(other.pk):
            foreign.refresh_from_db()
        self.assertEqual(foreign.status, StudentStatus.ACTIVE)


# --------------------------------------------------------------------------- #
# Bulk import — template download
# --------------------------------------------------------------------------- #
class BulkImportTemplateTests(StudentsTestCase):
    def setUp(self):
        super().setUp()
        self.url = reverse("students:import_template")

    def test_the_template_has_a_worksheet_per_active_class(self):
        self.sign_in_owner()
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], XLSX_CONTENT_TYPE)
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn(self.institution.code, response["Content-Disposition"])

        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ["JSS 1A", "JSS 1B"])
        first_sheet = workbook[workbook.sheetnames[0]]
        header = [cell.value for cell in first_sheet[1]]
        self.assertEqual(header, list(IMPORT_HEADERS))

    def test_the_template_needs_at_least_one_active_class(self):
        """No active class means no worksheet to generate — the view sends the
        user to class setup rather than hand back an empty workbook."""
        with self.in_school():
            Class.unscoped.filter(institution_id=self.institution.pk).update(
                status=ClassStatus.INACTIVE
            )
        self.sign_in_owner()
        response = self.client.get(self.url)
        self.assertRedirects(response, reverse("academic:class_list"))

    def test_a_bursar_cannot_download_the_template(self):
        """bulk_import is Owner/Administrator for view and manage alike."""
        self.sign_in_as(User.Role.BURSAR)
        self.assertEqual(self.client.get(self.url).status_code, 403)


# --------------------------------------------------------------------------- #
# Bulk import — upload + staging (nothing written to students yet)
# --------------------------------------------------------------------------- #
class BulkImportStagingTests(StudentsTestCase):
    def test_a_valid_upload_stages_rows_and_writes_no_students(self):
        self.sign_in_owner()
        response = self.upload(
            {"JSS 1A": [sheet_row("Amaka", "Eze"), sheet_row("Bola", "Okoro")]}
        )
        batch = self.latest_batch()
        self.assertRedirects(
            response, reverse("students:import_preview", args=[batch.pk])
        )

        with self.in_school():
            self.assertEqual(batch.status, BulkImportStatus.READY)
            rows = list(BulkImportRow.objects.filter(batch=batch))
            self.assertEqual(len(rows), 2)
            self.assertTrue(
                all(r.validation_status == RowValidationStatus.VALID for r in rows)
            )
            # The whole point of staging: the students table is still empty.
            self.assertEqual(Student.objects.count(), 0)

        preview = self.client.get(reverse("students:import_preview", args=[batch.pk]))
        self.assertEqual(preview.context["valid_count"], 2)
        self.assertEqual(preview.context["error_count"], 0)

    def test_rows_on_a_worksheet_that_matches_no_class_are_flagged(self):
        """A bad tab name fails every row on it, whatever the row contains —
        the mechanism that stops an import inventing a class."""
        self.sign_in_owner()
        self.upload({"Nonexistent Class": [sheet_row("Amaka", "Eze")]})
        batch = self.latest_batch()

        with self.in_school():
            row = BulkImportRow.objects.get(batch=batch)
        self.assertEqual(row.validation_status, RowValidationStatus.ERROR)
        self.assertIn("does not match an active class", row.error_reason)

    def test_a_row_missing_a_required_field_is_flagged(self):
        self.sign_in_owner()
        self.upload({"JSS 1A": [sheet_row("Amaka", "")]})  # no last name
        batch = self.latest_batch()

        with self.in_school():
            row = BulkImportRow.objects.get(batch=batch)
        self.assertEqual(row.validation_status, RowValidationStatus.ERROR)
        self.assertIn("Last Name", row.error_reason)

    def test_duplicates_within_the_file_and_against_existing_students(self):
        """Name + DOB keys a duplicate (no admission number exists at import
        time). Both a repeat inside the file and a clash with a student already
        on file are caught."""
        existing = self.enroll_a_student(self.jss1a, self.session, self.term)
        self.sign_in_owner()
        self.upload(
            {
                "JSS 1A": [
                    # Same name + DOB as the student already on file.
                    sheet_row(
                        existing.first_name, existing.last_name, dob="2012-04-02"
                    ),
                    # A pair identical to each other.
                    sheet_row("Ada", "Nnaji", dob="2013-01-01"),
                    sheet_row("Ada", "Nnaji", dob="2013-01-01"),
                ]
            }
        )
        batch = self.latest_batch()

        with self.in_school():
            reasons = list(
                BulkImportRow.objects.filter(
                    batch=batch, validation_status=RowValidationStatus.ERROR
                ).values_list("error_reason", flat=True)
            )
        self.assertTrue(any("already exists" in r for r in reasons), reasons)
        self.assertTrue(
            any("same name and date of birth" in r for r in reasons), reasons
        )

    def test_a_non_xlsx_upload_is_a_field_error_not_a_500(self):
        """clean_file refuses anything but .xlsx before the parser ever runs."""
        self.sign_in_owner()
        bad = SimpleUploadedFile("students.csv", b"a,b,c\n1,2,3\n", content_type="text/csv")
        response = self.client.post(
            reverse("students:import"), {"term": self.term.pk, "file": bad}
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("file", response.context["form"].errors)
        with self.in_school():
            self.assertEqual(BulkImportBatch.objects.count(), 0)

    def test_a_bursar_cannot_reach_the_upload(self):
        self.sign_in_as(User.Role.BURSAR)
        self.assertEqual(self.client.get(reverse("students:import")).status_code, 403)
        response = self.upload({"JSS 1A": [sheet_row("Amaka", "Eze")]})
        self.assertEqual(response.status_code, 403)
        with self.in_school():
            self.assertEqual(BulkImportBatch.objects.count(), 0)


# --------------------------------------------------------------------------- #
# Bulk import — commit (the single writer into students)
# --------------------------------------------------------------------------- #
class BulkImportCommitTests(StudentsTestCase):
    def test_committing_creates_students_one_audit_and_clears_staging(self):
        self.sign_in_owner()
        self.upload(
            {"JSS 1A": [sheet_row("Amaka", "Eze"), sheet_row("Bola", "Okoro")]}
        )
        batch = self.latest_batch()
        response = self.commit(batch, self.valid_row_pks(batch))
        self.assertRedirects(response, reverse("students:list"))

        with self.in_school():
            batch.refresh_from_db()
            self.assertEqual(batch.status, BulkImportStatus.COMMITTED)
            students = list(Student.objects.all())
            self.assertEqual(len(students), 2)
            for student in students:
                self.assertRegex(
                    student.admission_number,
                    rf"^{re.escape(self.institution.code)}-\d{{4}}-\d{{6}}$",
                )
                enrollment = StudentEnrollment.objects.get(student=student)
                self.assertEqual(enrollment.klass_id, self.jss1a.pk)
                self.assertEqual(enrollment.term_id, self.term.pk)
            # Admission numbers are distinct — the counter advanced per row.
            self.assertEqual(
                len({s.admission_number for s in students}), 2
            )
            # One summarising entry, not one per student.
            entry = AuditLog.objects.filter(action="student.bulk_imported").get()
            self.assertEqual(entry.detail["created"], 2)
            # Staging rows have served their purpose and are gone.
            self.assertEqual(BulkImportRow.objects.filter(batch=batch).count(), 0)

    def test_deselected_rows_are_not_imported(self):
        self.sign_in_owner()
        self.upload(
            {"JSS 1A": [sheet_row("Amaka", "Eze"), sheet_row("Bola", "Okoro")]}
        )
        batch = self.latest_batch()
        first_pk = self.valid_row_pks(batch)[0]  # Amaka, by insertion order
        self.commit(batch, [first_pk])

        with self.in_school():
            names = set(Student.objects.values_list("first_name", flat=True))
        self.assertEqual(names, {"Amaka"})

    def test_committing_with_nothing_selected_imports_nobody(self):
        self.sign_in_owner()
        self.upload({"JSS 1A": [sheet_row("Amaka", "Eze")]})
        batch = self.latest_batch()
        response = self.commit(batch, [])
        self.assertRedirects(response, reverse("students:list"))

        with self.in_school():
            batch.refresh_from_db()
            self.assertEqual(batch.status, BulkImportStatus.COMMITTED)
            self.assertEqual(Student.objects.count(), 0)

    def test_a_second_commit_does_not_import_the_batch_again(self):
        """Idempotent against a double submit — an already-committed batch is
        left untouched, so a browser back-and-resubmit cannot double the roll."""
        self.sign_in_owner()
        self.upload(
            {"JSS 1A": [sheet_row("Amaka", "Eze"), sheet_row("Bola", "Okoro")]}
        )
        batch = self.latest_batch()
        valid = self.valid_row_pks(batch)
        self.commit(batch, valid)
        # Second POST to the same preview — the rows are gone, but the guard is
        # the batch status, not their presence.
        second = self.commit(batch, valid)
        self.assertRedirects(second, reverse("students:list"))

        with self.in_school():
            self.assertEqual(Student.objects.count(), 2)
            self.assertEqual(
                AuditLog.objects.filter(action="student.bulk_imported").count(), 1
            )

    def test_error_rows_cannot_be_committed_even_if_posted(self):
        """The commit filters to valid rows in the service, so a hand-built POST
        that ticks an error row's id still cannot enrol it — the same DB-level,
        not template-level, guarantee the rest of the app relies on."""
        self.sign_in_owner()
        self.upload(
            {"JSS 1A": [sheet_row("Amaka", "Eze"), sheet_row("Broken", "")]}  # no last name
        )
        batch = self.latest_batch()
        self.commit(batch, self.all_row_pks(batch))  # try to commit everything

        with self.in_school():
            names = list(Student.objects.values_list("first_name", flat=True))
        self.assertEqual(names, ["Amaka"])

    def test_a_class_deactivated_after_staging_is_skipped_at_commit(self):
        """A row whose class was deactivated between preview and commit is
        skipped, not enrolled into an inactive class — the count reflects what
        was actually created."""
        self.sign_in_owner()
        self.upload({"JSS 1A": [sheet_row("Amaka", "Eze")]})
        batch = self.latest_batch()
        valid = self.valid_row_pks(batch)
        with self.in_school():
            Class.unscoped.filter(pk=self.jss1a.pk).update(status=ClassStatus.INACTIVE)
        self.commit(batch, valid)

        with self.in_school():
            self.assertEqual(Student.objects.count(), 0)
            entry = AuditLog.objects.filter(action="student.bulk_imported").get()
        self.assertEqual(entry.detail["created"], 0)

    def test_previewing_another_schools_batch_is_a_404(self):
        other = other_school()
        with institution_db_context(other.pk):
            o_session = Session.unscoped.create(
                institution=other,
                name="2026/2027",
                start_date=SESSION_START,
                end_date=SESSION_END,
                is_current=True,
            )
            o_term = Term.unscoped.create(
                institution=other,
                session=o_session,
                name="First Term",
                start_date=TERM_START,
                end_date=TERM_END,
                is_current=True,
            )
            foreign = BulkImportBatch.unscoped.create(
                institution=other,
                session=o_session,
                term=o_term,
                status=BulkImportStatus.READY,
            )
        self.sign_in_owner()
        response = self.client.get(
            reverse("students:import_preview", args=[foreign.pk])
        )
        self.assertEqual(response.status_code, 404)


# --------------------------------------------------------------------------- #
# Anonymous access
# --------------------------------------------------------------------------- #
class AnonymousAccessTests(TestCase):
    """No student screen is public — every one bounces an anonymous caller to
    login rather than 404ing, which would tell a stranger which ids exist."""

    def test_every_student_url_redirects_anonymous_to_login(self):
        for name, args in (
            ("students:list", ()),
            ("students:add", ()),
            ("students:import", ()),
            ("students:import_template", ()),
            ("students:import_preview", (1,)),
            ("students:detail", (1,)),
            ("students:edit", (1,)),
            ("students:delete", (1,)),
        ):
            with self.subTest(url=name):
                response = self.client.get(reverse(name, args=args))
                self.assertEqual(response.status_code, 302)
                self.assertIn(reverse("accounts:login"), response["Location"])


class StudentDeleteTests(ApprovedSchoolTestCase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        with institution_db_context(cls.institution.pk):
            cls.owner = User.objects.get(email=cls.OWNER_EMAIL)
            cls.session = Session.unscoped.create(
                institution=cls.institution,
                name="2026/2027",
                start_date=SESSION_START,
                end_date=SESSION_END,
                is_current=True,
            )
            cls.term = Term.unscoped.create(
                institution=cls.institution,
                session=cls.session,
                name="First Term",
                start_date=TERM_START,
                end_date=TERM_END,
                is_current=True,
            )
            cls.klass = Class.unscoped.create(
                institution=cls.institution, name="JSS 1", order=1
            )

    def _create_test_student(self, first_name="Chidi", last_name="Eze"):
        with self.in_school():
            from students.services import create_student
            student = Student(
                first_name=first_name,
                last_name=last_name,
                gender=Gender.MALE,
                date_of_birth=datetime.date(2014, 5, 12),
                guardian_name="Emeka Eze",
                guardian_phone="08031234567",
            )
            return create_student(
                institution=self.institution,
                student=student,
                klass=self.klass,
                session=self.session,
                term=self.term,
                actor=self.owner,
            )

    def test_delete_student_purges_all_records(self):
        student = self._create_test_student()
        with self.in_school():
            self.assertEqual(Student.objects.filter(pk=student.pk).count(), 1)
            self.assertEqual(StudentEnrollment.objects.filter(student=student).count(), 1)

            res = delete_student(
                student=student,
                keep_financial_records=False,
                actor=self.owner,
            )
            self.assertEqual(res["action"], "deleted")
            self.assertEqual(Student.objects.filter(pk=student.pk).count(), 0)
            self.assertEqual(StudentEnrollment.objects.filter(student_id=student.pk).count(), 0)

            log = AuditLog.objects.filter(action="student.deleted").first()
            self.assertIsNotNone(log)

    def test_delete_student_with_keep_financial_records_archives(self):
        student = self._create_test_student()
        with self.in_school():
            res = delete_student(
                student=student,
                keep_financial_records=True,
                actor=self.owner,
            )
            self.assertEqual(res["action"], "archived")
            student.refresh_from_db()
            self.assertEqual(student.status, StudentStatus.INACTIVE)

    def test_student_delete_view_get_and_post(self):
        student = self._create_test_student()
        self.sign_in_owner()
        url = reverse("students:delete", args=[student.pk])
        
        # GET confirmation page
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, student.full_name)

        # POST delete
        response = self.client.post(url, {"keep_records": "off"})
        self.assertRedirects(response, reverse("students:list"))
        with self.in_school():
            self.assertEqual(Student.objects.filter(pk=student.pk).count(), 0)

